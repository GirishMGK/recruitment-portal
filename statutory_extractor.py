import streamlit as st
import pandas as pd
import re
import fitz
from io import BytesIO
import zipfile
import os
import datetime

st.set_page_config(page_title="Statutory Compliance Extractor", layout="wide")
st.title("📋 Statutory Compliance Extractor")
st.caption("ESI · PT · TDS (ITNS 281) · GSTR-1 · GSTR-3B")

# ================= SESSION STATE =================
TYPES = ["esi", "pt", "tds", "gstr1", "gstr3b"]
for t in TYPES:
    if t not in st.session_state:
        st.session_state[t] = []
if "failed" not in st.session_state:
    st.session_state.failed = []
if "recon" not in st.session_state:
    st.session_state.recon = []

# ================= UTILS =================

def read_pdf_text(file_bytes):
    pdf = fitz.open(stream=file_bytes, filetype="pdf")
    return "\n".join(page.get_text(sort=True) for page in pdf)

def g(pattern, text, flags=re.I):
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else ""

def normalize_period(raw):
    if not raw:
        return ""
    raw = str(raw).strip()
    cleaned = re.sub(r'\s*[-–]\s*', '-', raw)
    for candidate in [cleaned, cleaned.title(), raw, raw.title()]:
        for fmt in ["%b-%y", "%b-%Y", "%B-%Y", "%B-%y", "%m-%Y", "%B"]:
            try:
                dt = datetime.datetime.strptime(candidate, fmt)
                if fmt == "%B":
                    return candidate.title()
                return dt.strftime("%b-%Y")
            except:
                pass
    return raw

_MONTH_NUM = {
    1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
    7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"
}

def period_from_filename(file_name):
    """Extract (month_name, FY) from filenames like GSTR1_GST_042025.pdf.
    Looks for the first MMYYYY or MMYY pattern in the base file name.
    Returns ("", "") if nothing found."""
    base = os.path.splitext(os.path.basename(file_name))[0]
    m = re.search(r'[_\-\s](\d{2})(\d{4})(?:[_\-\s]|$)', base)
    if not m:
        m = re.search(r'(\d{2})(\d{4})', base)
    if m:
        mm, yyyy = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12 and 2000 <= yyyy <= 2099:
            month_name = _MONTH_NUM[mm]
            fy = f"{yyyy}-{str(yyyy+1)[2:]}" if mm >= 4 else f"{yyyy-1}-{str(yyyy)[2:]}"
            return month_name, fy
    return "", ""


# ================= DETECTOR =================

def detect_type(text):
    # ── 1. TDS ────────────────────────────────────────────────────────────────
    if re.search(r"ITNS[\s\w.:]*281", text, re.I):
        return "TDS"
    # ── GSTR-2B — must be checked BEFORE GSTR-3B because GSTR-2B PDFs
    #    mention "GSTR-3B" in their header/instructions ──────────────────────
    if re.search(r"FORM\s*GSTR-2B|Form GSTR-2B|GSTR\s*[-–]\s*2B", text):
        return "GSTR2B"
    # ── 4. GSTR-3B ───────────────────────────────────────────────────────────
    if re.search(r"FORM\s*GSTR-3B|Form GSTR-3B", text):
        return "GSTR3B"
    # ── 5. GSTR-1 ────────────────────────────────────────────────────────────
    if re.search(r"FORM\s*GSTR-1|Details of outward supplies", text):
        return "GSTR1"
    # ── 6. PT ─────────────────────────────────────────────────────────────────
    if re.search(r"FORM\s*5-A|PROFESSIONAL TAX RETURNS", text, re.I):
        return "PT"
    # ── 7. ESI ───────────────────────────────────────────────────────────────
    if re.search(r"Employer.s Code No|esic\.in|Employees.?\s*State Insurance", text, re.I):
        return "ESI"
    return None

# ================= EXTRACTOR: ESI =================

def extract_esi(file_name, text):
    period_raw = g(r"Challan Period\s*:\s*([^\n]+)", text)
    submitted  = (
        g(r"Challan Submitted Date\s*:\s*([^\n]+)", text) or
        g(r"Challan Submitted Date\s+(\d{2}-\d{2}-\d{4}[^\n]*)", text)
    )
    return {
        "File Name":              file_name,
        "Employer Name":          g(r"Employer.s Name\s*:\s*([^\n]+)", text),
        "Employer Code No":       g(r"Employer.s Code No\s*:\s*([^\n]+)", text),
        "Challan Period":         normalize_period(period_raw),
        "Challan Number":         g(r"Challan Number\s*[:\s]+([^\n]+)", text),
        "Challan Created Date":   g(r"Challan Created Date\s+([^\n]+)", text),
        "Challan Submitted Date": submitted,
        "Amount Paid":            g(r"Amount Paid\s*:\s*([\d.,]+)", text),
        "Transaction Number":     g(r"Transaction Number\s*:\s*([^\n]+)", text),
        "Transaction Status":     g(r"Transaction status\s*:\s*([^\n]+)", text),
    }

# ================= EXTRACTOR: PT =================

def extract_pt(file_name, text):
    period_raw = (
        g(r"ending on\s*:\s*([A-Za-z]+\s*[-–]\s*\d{4})", text) or
        g(r"for the month.*?:\s*([A-Za-z]+\s*[-–]\s*\d{4})", text, re.I | re.S)
    )
    grand_total  = g(r"Grand Total\s+([\d,]+)", text) or g(r"Grand Total\s*:\s*([\d,]+)", text)
    pay_date     = (
        g(r"(\d{2}-\d{2}-\d{4})\s+Khajane", text) or
        g(r"ePayment\s*\d+\s+\S+\s+(\d{2}[-/]\d{2}[-/]\d{4})", text)
    )
    amount_paid  = (
        g(r"Khajane\s*\nII\n([\d,]+)", text) or
        g(r"Khajane\s*II\s*([\d,]+)", text) or grand_total
    )
    return {
        "File Name":     file_name,
        "Client Name":   g(r"Trade Name\s*:\s*(.+)", text) or g(r"Name of the Employer\s*:\s*(.+)", text),
        "Reg. Cert. No": g(r"Registration Certificate Number\s*:\s*([^\n]+)", text),
        "Period":        normalize_period(period_raw),
        "Return Date":   g(r"Return Date\s*:\s*(\d{2}-\d{2}-\d{4})", text),
        "Grand Total":   grand_total,
        "Payment Date":  pay_date,
        "Amount Paid":   amount_paid,
    }

# ================= EXTRACTOR: TDS =================

def extract_tds(file_name, text):
    client = (
        g(r"Name of Deductor\s*:\s*([^\n]+)", text) or
        g(r"Name of Tax Payer\s*:\s*([^\n]+)", text) or
        g(r"(?m)^Name\s*:\s*([^\n]+)", text) or
        g(r"\bName\s*:\s*([A-Z][A-Z ]+)", text)
    )
    amount    = g(r"Amount\s*\(in Rs\.\)\s*:\s*[^\d]*([\d,]+)", text)
    tax       = g(r"\bA\b\s+Tax\s+[^\d]*([\d,]+)", text)
    surcharge = g(r"\bB\b\s+Surcharge\s+[^\d]*([\d,]+)", text)
    cess      = g(r"\bC\b\s+Cess\s+[^\d]*([\d,]+)", text)
    interest  = g(r"\bD\b\s+Interest\s+[^\d]*([\d,]+)", text)
    penalty   = g(r"\bE\b\s+Penalty\s+[^\d]*([\d,]+)", text)
    fee_234e  = g(r"(?:F\s+)?Fee under section 234E\s+[^\d]*([\d,]+)", text)
    return {
        "File Name":         file_name,
        "Client Name":       client,
        "TAN":               g(r"TAN\s*:\s*([A-Z]{4}\d{5}[A-Z])", text),
        "Assessment Year":   g(r"Assessment Year\s*:\s*([\d-]+)", text),
        "Financial Year":    g(r"Financial Year\s*:\s*([\d-]+)", text),
        "Major Head":        g(r"Major Head\s*:\s*([^\n]+)", text),
        "Minor Head":        g(r"Minor Head\s*:\s*([^\n]+)", text),
        "Nature of Payment": g(r"Nature of Payment\s*:\s*([^\n]+)", text),
        "Amount (Rs)":       amount,
        "Date of Deposit":   g(r"Date of Deposit\s*:\s*([\w-]+)", text),
        "CIN":               g(r"CIN\s*:\s*([A-Z0-9]+)", text),
        "Bank Name":         g(r"Bank Name\s*:\s*([^\n]+)", text),
        "Challan No":        g(r"Challan No\s*:\s*(\d+)", text),
        "Mode of Payment":   g(r"Mode of Payment\s*:\s*([^\n]+)", text),
        "Tax":               tax,
        "Surcharge":         surcharge,
        "Cess":              cess,
        "Interest":          interest,
        "Penalty":           penalty,
        "Fee u/s 234E":      fee_234e,
    }

# ================= EXTRACTOR: GSTR-1 =================

def extract_gstr1(file_name, text):
    gstin  = g(r"(?:1\s+)?GSTIN\s+(\w{15})", text)
    client = (g(r"Legal name of the registered person\s+([^\n]+)", text) or
              g(r"Trade name if any\s+([^\n]+)", text))
    fin_yr = g(r"Financial year\s+([\d\-]+)", text)
    period = g(r"Tax period\s+([A-Za-z]+)", text)
    if period and period.title() not in _MONTH_NUM.values():
        period = ""
    fn_period, fn_fy = period_from_filename(file_name)
    if not period:
        period = fn_period
    if not fin_yr:
        fin_yr = fn_fy
    arn    = (g(r"\(c\)\s*ARN\s+([A-Z0-9]+)", text) or g(r"\bARN\b\s+([A-Z0-9]+)", text))
    arn_date = (g(r"\(d\)\s*ARN date\s+([\d/]+)", text) or g(r"ARN date\s+([\d/]+)", text))

    def block_row5(block_text):
        nm = re.search(
            r"Total\s+\d+\s+Invoice\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
            block_text, re.S)
        if nm:
            return tuple(v.replace(",","") for v in nm.groups())
        return ("","","","","")

    def block_count(block_text, doc_type="Invoice"):
        cm = re.search(rf"Total\s+(\d+)\s+{doc_type}", block_text, re.I)
        return cm.group(1) if cm else ""

    t4a_blk = re.search(r"4A[^\n]*B2B Regular(.*?)(?=4B|5\s+-\s+Taxable|$)", text, re.S|re.I)
    t4a = block_row5(t4a_blk.group(1)) if t4a_blk else ("","","","","")
    t4a_n = block_count(t4a_blk.group(1)) if t4a_blk else ""

    t4b_blk = re.search(r"4B[^\n]*Reverse charge(.*?)(?=5\s+-\s+Taxable|6A|$)", text, re.S|re.I)
    t4b = block_row5(t4b_blk.group(1)) if t4b_blk else ("","","","","")
    t4b_n = block_count(t4b_blk.group(1)) if t4b_blk else ""

    t7_blk = re.search(r"7\s*-\s*Taxable supplies.*?B2CS(.*?)(?=8\s*-\s*Nil|$)", text, re.S|re.I)
    t7, t7_n = ("","","","",""), ""
    if t7_blk:
        blk  = t7_blk.group(1)
        t7_n = block_count(blk, "Net Value")
        nm = re.search(
            r"Total\s+\d+\s+Net Value\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
            blk, re.S)
        if nm:
            t7 = tuple(v.replace(",","") for v in nm.groups())

    def table_totals(blk):
        """Extract 5-column totals row from a GSTR-1 table block."""
        nm = re.search(
            r"Total.*?([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
            blk, re.S)
        return tuple(v.replace(",","") for v in nm.groups()) if nm else ("","","","","")

    t11a_blk = re.search(
        r"11A[^\n]*(?:Advance received|invoice has not been issued)(.*?)(?=11B|12\s*[-–]|$)",
        text, re.S|re.I)
    t11a, t11a_n = ("","","","",""), ""
    if t11a_blk:
        blk = t11a_blk.group(1)
        cm  = re.search(r"Total\s+(\d+)\s+NA", blk, re.I)
        t11a_n = cm.group(1) if cm else (re.search(r"Total\s+(\d+)", blk) or type("", (), {"group": lambda *_: ""})()).group(1)
        t11a = table_totals(blk)

    t11b_blk = re.search(
        r"11B[^\n]*(?:Advance adjusted|adjusted against)(.*?)(?=12\s*[-–]|$)",
        text, re.S|re.I)
    t11b, t11b_n = ("","","","",""), ""
    if t11b_blk:
        blk = t11b_blk.group(1)
        cm  = re.search(r"Total\s+(\d+)\s+NA", blk, re.I)
        t11b_n = cm.group(1) if cm else (re.search(r"Total\s+(\d+)", blk) or type("", (), {"group": lambda *_: ""})()).group(1)
        t11b = table_totals(blk)

    t12_blk = re.search(r"12\s*-\s*HSN.wise summary(.*?)(?=13\s*-|$)", text, re.S|re.I)
    t12, t12_n = ("","","","",""), ""
    if t12_blk:
        blk  = t12_blk.group(1)
        cm   = re.search(r"Total\s+(\d+)\s+NA", blk, re.I)
        t12_n = cm.group(1) if cm else ""
        nm = re.search(
            r"Total\s+\d+\s+NA\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
            blk, re.S)
        if nm:
            t12 = tuple(v.replace(",","") for v in nm.groups())

    tl_m = re.search(
        r"Total Liability.*?\b([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
        text, re.S|re.I)
    tl = tuple(v.replace(",","") for v in tl_m.groups()) if tl_m else ("","","","","")

    return {
        "File Name": file_name, "Client Name": client, "GSTIN": gstin,
        "Financial Year": fin_yr, "Tax Period": period, "ARN": arn, "ARN Date": arn_date,
        "4A No of Invoices": t4a_n, "4A Taxable Value": t4a[0], "4A IGST": t4a[1],
        "4A CGST": t4a[2], "4A SGST": t4a[3], "4A Cess": t4a[4],
        "4B No of Invoices": t4b_n, "4B Taxable Value": t4b[0], "4B IGST": t4b[1],
        "4B CGST": t4b[2], "4B SGST": t4b[3], "4B Cess": t4b[4],
        "7 No of Records": t7_n, "7 Taxable Value": t7[0], "7 IGST": t7[1],
        "7 CGST": t7[2], "7 SGST": t7[3], "7 Cess": t7[4],
        "11A No of Records": t11a_n, "11A Taxable Value": t11a[0], "11A IGST": t11a[1],
        "11A CGST": t11a[2], "11A SGST": t11a[3], "11A Cess": t11a[4],
        "11B No of Records": t11b_n, "11B Taxable Value": t11b[0], "11B IGST": t11b[1],
        "11B CGST": t11b[2], "11B SGST": t11b[3], "11B Cess": t11b[4],
        "12 No of HSN": t12_n, "12 Taxable Value": t12[0], "12 IGST": t12[1],
        "12 CGST": t12[2], "12 SGST": t12[3], "12 Cess": t12[4],
        "TL Taxable Value": tl[0], "TL IGST": tl[1], "TL CGST": tl[2],
        "TL SGST": tl[3], "TL Cess": tl[4],
    }

# ================= EXTRACTOR: GSTR-3B =================

def extract_gstr3b(file_name, text):
    gstin  = (g(r"GSTIN of the supplier\s+(\w{15})", text) or g(r"GSTIN\s+(\w{15})", text))
    client = (g(r"2\(a\)\.\s*Legal name of the registered person\s+([^\n]+)", text) or
              g(r"Legal name of the registered person\s+([^\n]+)", text))
    # Try several patterns to handle different PDF text layouts; fallback to filename
    fin_yr = (g(r"^Year\s+([\d\-]+)", text, re.M) or
              g(r"(?m)^\s*2\s*\(b\)[.\s]*Year\s+([\d\-]+)", text) or
              g(r"\bYear\s+(20\d{2}[-–]\d{2,4})\b", text))
    period = (g(r"^Period\s+([A-Za-z]+)", text, re.M) or
              g(r"(?m)^\s*2\s*\(c\)[.\s]*Period\s+([A-Za-z]+)", text) or
              g(r"\bPeriod\s+([A-Z][a-z]+)\b", text))
    # Validate: reject if "period" captured a non-month word like "Integrated"
    if period and period.title() not in _MONTH_NUM.values():
        period = ""
    fn_period, fn_fy = period_from_filename(file_name)
    if not period:
        period = fn_period
    if not fin_yr:
        fin_yr = fn_fy
    arn    = (g(r"2\(c\)\.\s*ARN\s+([A-Z0-9]+)", text) or g(r"\bARN\b\s+([A-Z0-9]+)", text))
    arn_date = g(r"Date of ARN\s+([\d/]+)", text)

    def row5(marker):
        pat = (re.escape(marker) +
               r"[^0-9]*?([\d,]+\.\d{2})[^0-9]*?([\d,]+\.\d{2})"
               r"[^0-9]*?([\d,]+\.\d{2})[^0-9]*?([\d,]+\.\d{2})"
               r"[^0-9]*?([\d,]+\.\d{2})")
        m = re.search(pat, text, re.S | re.I)
        return tuple(v.replace(",","") for v in m.groups()) if m else ("","","","","")

    def itc_row(pattern):
        m = re.search(
            pattern +
            r"[^0-9]*?([\d,]+\.\d{2})[^0-9]*?([\d,]+\.\d{2})"
            r"[^0-9]*?([\d,]+\.\d{2})[^0-9]*?([\d,]+\.\d{2})",
            text, re.S | re.I)
        return tuple(v.replace(",","") for v in m.groups()) if m else ("","","","")

    t31a  = row5("(a) Outward taxable supplies (other than zero rated")
    t31d  = row5("(d) Inward supplies (liable to reverse charge)")
    itc_c = itc_row(r"C\.\s*Net ITC available")

    sec_a = re.search(r"\(A\)\s*Other than reverse charge(.*?)(?=\(B\)\s*Reverse)", text, re.S|re.I)
    sec_b = re.search(r"\(B\)\s*Reverse charge(.*?)(?=Breakup|Verif|$)", text, re.S|re.I)

    def get_61(label_pat, src):
        if not src:
            return ("","","")
        pat = (label_pat +
               r"\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})"
               r"((?:\s+[-\d,]+\.?\d*)*?)\s+([\d,]+\.\d{2})")
        m = re.search(pat, src, re.S|re.I)
        if m:
            return (m.group(1).replace(",",""), m.group(3).replace(",",""), m.group(5).replace(",",""))
        return ("","","")

    sa = sec_a.group(1) if sec_a else ""
    sb = sec_b.group(1) if sec_b else ""

    igst_a_pay, igst_a_net, igst_a_cash = get_61(r"Integrated\s*\n?\s*tax", sa)
    cgst_a_pay, cgst_a_net, cgst_a_cash = get_61(r"Central\s*\n?\s*tax",    sa)
    sgst_a_pay, sgst_a_net, sgst_a_cash = get_61(r"State/UT\s*\n?\s*tax",   sa)
    igst_b_pay, igst_b_net, igst_b_cash = get_61(r"Integrated\s*\n?\s*tax", sb)
    cgst_b_pay, cgst_b_net, cgst_b_cash = get_61(r"Central\s*\n?\s*tax",    sb)
    sgst_b_pay, sgst_b_net, sgst_b_cash = get_61(r"State/UT\s*\n?\s*tax",   sb)

    bp = re.search(
        r"(?:January|February|March|April|May|June|July|August|September|October|November|December)"
        r"\s+\d{4}\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
        text, re.I)

    return {
        "File Name": file_name, "Client Name": client, "GSTIN": gstin,
        "Financial Year": fin_yr, "Tax Period": period, "ARN": arn, "ARN Date": arn_date,
        "3.1(a) Taxable Value": t31a[0], "3.1(a) IGST": t31a[1], "3.1(a) CGST": t31a[2],
        "3.1(a) SGST": t31a[3], "3.1(a) Cess": t31a[4],
        "3.1(d) RC Value": t31d[0], "3.1(d) RC IGST": t31d[1],
        "3.1(d) RC CGST": t31d[2], "3.1(d) RC SGST": t31d[3],
        "4(C) Net ITC IGST": itc_c[0], "4(C) Net ITC CGST": itc_c[1], "4(C) Net ITC SGST": itc_c[2],
        "6.1(A) IGST Payable": igst_a_pay, "6.1(A) IGST Net": igst_a_net, "6.1(A) IGST Cash": igst_a_cash,
        "6.1(A) CGST Payable": cgst_a_pay, "6.1(A) CGST Net": cgst_a_net, "6.1(A) CGST Cash": cgst_a_cash,
        "6.1(A) SGST Payable": sgst_a_pay, "6.1(A) SGST Net": sgst_a_net, "6.1(A) SGST Cash": sgst_a_cash,
        "6.1(B) IGST Payable": igst_b_pay, "6.1(B) IGST Cash": igst_b_cash,
        "6.1(B) CGST Payable": cgst_b_pay, "6.1(B) CGST Cash": cgst_b_cash,
        "6.1(B) SGST Payable": sgst_b_pay, "6.1(B) SGST Cash": sgst_b_cash,
        "Breakup IGST": bp.group(1).replace(",","") if bp else "",
        "Breakup CGST": bp.group(2).replace(",","") if bp else "",
        "Breakup SGST": bp.group(3).replace(",","") if bp else "",
        "Breakup Cess": bp.group(4).replace(",","") if bp else "",
    }

# ================= RECONCILIATION: GSTR-1 vs GSTR-3B =================

def compute_recon(gstr1_rows, gstr3b_rows):
    """Match GSTR-1 and GSTR-3B records by GSTIN + Tax Period and diff key amounts."""

    def to_f(val):
        try:
            return float(str(val).replace(",", "")) if val not in ("", None) else 0.0
        except ValueError:
            return 0.0

    # Index 3B records by (GSTIN, normalised period)
    idx3b = {}
    for row in gstr3b_rows:
        key = (str(row.get("GSTIN", "")).strip().upper(),
               normalize_period(row.get("Tax Period", "")))
        idx3b[key] = row

    results = []
    matched_3b_keys = set()

    for r1 in gstr1_rows:
        gstin  = str(r1.get("GSTIN", "")).strip().upper()
        period = normalize_period(r1.get("Tax Period", ""))
        key    = (gstin, period)
        r3b    = idx3b.get(key)

        # GSTR-1 values from Total Liability
        g1_tv   = to_f(r1.get("TL Taxable Value"))
        g1_igst = to_f(r1.get("TL IGST"))
        g1_cgst = to_f(r1.get("TL CGST"))
        g1_sgst = to_f(r1.get("TL SGST"))
        g1_tax  = g1_igst + g1_cgst + g1_sgst

        # GSTR-3B values from section 3.1(a)
        g3_tv   = to_f(r3b.get("3.1(a) Taxable Value")) if r3b else 0.0
        g3_igst = to_f(r3b.get("3.1(a) IGST"))          if r3b else 0.0
        g3_cgst = to_f(r3b.get("3.1(a) CGST"))          if r3b else 0.0
        g3_sgst = to_f(r3b.get("3.1(a) SGST"))          if r3b else 0.0
        g3_tax  = g3_igst + g3_cgst + g3_sgst

        diff_tv   = round(g1_tv   - g3_tv,   2)
        diff_igst = round(g1_igst - g3_igst, 2)
        diff_cgst = round(g1_cgst - g3_cgst, 2)
        diff_sgst = round(g1_sgst - g3_sgst, 2)
        diff_tax  = round(g1_tax  - g3_tax,  2)

        TOLERANCE = 1.0   # ₹1 tolerance for rounding differences
        if r3b is None:
            status = "Only in GSTR-1"
        elif abs(diff_tv) <= TOLERANCE and abs(diff_tax) <= TOLERANCE:
            status = "Matched"
        else:
            status = "Mismatch"

        if r3b:
            matched_3b_keys.add(key)

        results.append({
            "GSTIN":                  gstin,
            "Client Name":            r1.get("Client Name", ""),
            "Financial Year":         r1.get("Financial Year", ""),
            "Tax Period":             period,
            "GSTR-1 File":            r1.get("File Name", ""),
            "GSTR-3B File":           r3b.get("File Name", "") if r3b else "",
            "GSTR-1 Taxable Value":   g1_tv,
            "GSTR-3B Taxable Value":  g3_tv,
            "Diff Taxable Value":     diff_tv,
            "GSTR-1 IGST":            g1_igst,
            "GSTR-3B IGST":           g3_igst,
            "Diff IGST":              diff_igst,
            "GSTR-1 CGST":            g1_cgst,
            "GSTR-3B CGST":           g3_cgst,
            "Diff CGST":              diff_cgst,
            "GSTR-1 SGST":            g1_sgst,
            "GSTR-3B SGST":           g3_sgst,
            "Diff SGST":              diff_sgst,
            "GSTR-1 Total Tax":       g1_tax,
            "GSTR-3B Total Tax":      g3_tax,
            "Diff Total Tax":         diff_tax,
            "Status":                 status,
        })

    # Add 3B records with no matching GSTR-1
    for key, r3b in idx3b.items():
        if key in matched_3b_keys:
            continue
        g3_tv   = to_f(r3b.get("3.1(a) Taxable Value"))
        g3_igst = to_f(r3b.get("3.1(a) IGST"))
        g3_cgst = to_f(r3b.get("3.1(a) CGST"))
        g3_sgst = to_f(r3b.get("3.1(a) SGST"))
        g3_tax  = g3_igst + g3_cgst + g3_sgst
        results.append({
            "GSTIN":                  key[0],
            "Client Name":            r3b.get("Client Name", ""),
            "Financial Year":         r3b.get("Financial Year", ""),
            "Tax Period":             key[1],
            "GSTR-1 File":            "",
            "GSTR-3B File":           r3b.get("File Name", ""),
            "GSTR-1 Taxable Value":   0.0,
            "GSTR-3B Taxable Value":  g3_tv,
            "Diff Taxable Value":     round(0.0 - g3_tv, 2),
            "GSTR-1 IGST":            0.0,
            "GSTR-3B IGST":           g3_igst,
            "Diff IGST":              round(-g3_igst, 2),
            "GSTR-1 CGST":            0.0,
            "GSTR-3B CGST":           g3_cgst,
            "Diff CGST":              round(-g3_cgst, 2),
            "GSTR-1 SGST":            0.0,
            "GSTR-3B SGST":           g3_sgst,
            "Diff SGST":              round(-g3_sgst, 2),
            "GSTR-1 Total Tax":       0.0,
            "GSTR-3B Total Tax":      g3_tax,
            "Diff Total Tax":         round(-g3_tax, 2),
            "Status":                 "Only in GSTR-3B",
        })

    return results


# ================= PROCESS SINGLE PDF =================

def process_pdf(file_name, file_bytes):
    import traceback
    try:
        text     = read_pdf_text(file_bytes)

        # Guard: scanned / image PDFs return almost no selectable text.
        # Without this check the extractor silently produces empty rows or
        # crashes, and the file ends up in the failed list with a cryptic error.
        if len(text.strip()) < 50:
            st.session_state.failed.append(
                f"{file_name} — Skipped: PDF appears to be image-only / scanned "
                f"(no selectable text found). Use OCR before uploading."
            )
            return

        doc_type = detect_type(text)

        if doc_type == "GSTR2B":
            st.session_state.failed.append(
                f"{file_name} — ℹ️ GSTR-2B (skipped)"
            )
        elif doc_type in ("ESI","PT","TDS","GSTR1","GSTR3B"):
            dispatch = {
                "ESI":    ("esi",    extract_esi),
                "PT":     ("pt",     extract_pt),
                "TDS":    ("tds",    extract_tds),
                "GSTR1":  ("gstr1",  extract_gstr1),
                "GSTR3B": ("gstr3b", extract_gstr3b),
            }
            key, fn = dispatch[doc_type]
            st.session_state[key].append(fn(file_name, text))
        else:
            # Show the first 200 chars of extracted text so the user can
            # diagnose why the document wasn't recognised.
            preview = text.strip()[:200].replace("\n", " ")
            st.session_state.failed.append(
                f"{file_name} — Unrecognised type. "
                f"Text preview: «{preview}»"
            )
    except Exception as e:
        tb = traceback.format_exc().splitlines()
        # Keep only the last 4 lines of the traceback so it's readable in the UI
        short_tb = " | ".join(line.strip() for line in tb[-4:] if line.strip())
        st.session_state.failed.append(
            f"{file_name} — {type(e).__name__}: {e} — [{short_tb}]"
        )

# ================= INPUT UI =================

st.markdown("---")
st.subheader("📂 Step 1 — Upload Files")

col1, col2, col3 = st.columns(3)
with col1:
    pdf_files = st.file_uploader("🗂 Upload PDFs", type="pdf", accept_multiple_files=True, key="pdf_uploader")
with col2:
    zip_files = st.file_uploader("🗜 Upload ZIP files", type="zip", accept_multiple_files=True, key="zip_uploader")
with col3:
    folder_paths_raw = st.text_area(
        "📁 Folder paths (one per line)",
        placeholder="D:\\Challans\\Jan2025\nD:\\Challans\\Feb2025",
        height=100,
    )

st.markdown("---")
col_btn1, col_btn2, col_btn3 = st.columns(3)
with col_btn1:
    extract_btn = st.button("🚀 Extract All", use_container_width=True, type="primary")
with col_btn2:
    if st.button("🗑️ Clear Results", use_container_width=True):
        for t in TYPES:
            st.session_state[t] = []
        st.session_state.failed = []
        st.session_state.recon  = []
        st.success("✅ Results cleared — upload new files and click Extract All.")
with col_btn3:
    if st.button("🔄 Refresh / Reset All", use_container_width=True):
        for t in TYPES:
            st.session_state[t] = []
        st.session_state.failed = []
        st.session_state.recon  = []
        for k in ["pdf_uploader", "zip_uploader"]:
            if k in st.session_state:
                del st.session_state[k]
        st.rerun()

if extract_btn:
    for t in TYPES:
        st.session_state[t] = []
    st.session_state.failed = []
    st.session_state.recon  = []

    direct_pdfs = [(f.name, f.read()) for f in (pdf_files or [])]
    zip_pdfs = []
    for zf in (zip_files or []):
        with zipfile.ZipFile(zf) as z:
            for name in z.namelist():
                if name.lower().endswith(".pdf"):
                    zip_pdfs.append((os.path.basename(name), z.read(name)))
    folder_pdfs = []
    if folder_paths_raw:
        for folder in folder_paths_raw.strip().splitlines():
            folder = folder.strip()
            if not folder:
                continue
            if not os.path.isdir(folder):
                st.session_state.failed.append(f"Folder not found: {folder}")
                continue
            for root, _, files in os.walk(folder):
                for fn in files:
                    if fn.lower().endswith(".pdf"):
                        path = os.path.join(root, fn)
                        with open(path, "rb") as fp:
                            folder_pdfs.append((fn, fp.read()))

    all_pdfs = direct_pdfs + zip_pdfs + folder_pdfs
    if not all_pdfs:
        st.warning("⚠️ No PDF files found.")
    else:
        with st.spinner(f"Processing {len(all_pdfs)} file(s)..."):
            bar = st.progress(0)
            for i, (name, data) in enumerate(all_pdfs):
                process_pdf(name, data)
                bar.progress((i + 1) / len(all_pdfs))

        # Compute reconciliation after all PDFs are processed
        if st.session_state.gstr1 or st.session_state.gstr3b:
            st.session_state.recon = compute_recon(
                st.session_state.gstr1, st.session_state.gstr3b
            )

        counts = {t: len(st.session_state[t]) for t in TYPES}
        st.success(
            f"✅ Done — "
            f"**{counts['esi']}** ESI · "
            f"**{counts['pt']}** PT · "
            f"**{counts['tds']}** TDS · "
            f"**{counts['gstr1']}** GSTR-1 · "
            f"**{counts['gstr3b']}** GSTR-3B · "
            f"**{len(st.session_state.recon)}** reconciliation rows · "
            f"**{sum(1 for f in st.session_state.failed if 'GSTR-2B (skipped)' not in f and 'image-only' not in f)}** failed"
        )

# ================= FAILED FILES =================
if st.session_state.failed:
    # Categorise failures so the user can act on them
    gstr2b_skipped = [f for f in st.session_state.failed if "GSTR-2B (skipped)" in f]
    image_pdfs     = [f for f in st.session_state.failed if "image-only" in f]
    unrecognised   = [f for f in st.session_state.failed if "Unrecognised type" in f]
    errors         = [f for f in st.session_state.failed
                      if f not in gstr2b_skipped and f not in image_pdfs
                      and f not in unrecognised]

    # Only count actionable issues (errors + unrecognised) as "failed" in the header
    actionable = len(errors) + len(unrecognised)
    header_parts = []
    if errors:         header_parts.append(f"{len(errors)} error(s)")
    if unrecognised:   header_parts.append(f"{len(unrecognised)} unrecognised")
    if image_pdfs:     header_parts.append(f"{len(image_pdfs)} scanned PDF(s)")
    if gstr2b_skipped: header_parts.append(f"{len(gstr2b_skipped)} GSTR-2B skipped")

    expander_icon = "⚠️" if actionable else "ℹ️"
    with st.expander(f"{expander_icon} {' · '.join(header_parts)}  ▸ expand for details"):
        if errors:
            st.markdown("**🔴 Extraction errors** (bug or corrupted PDF):")
            for msg in errors:
                st.code(msg, language="")

        if unrecognised:
            st.markdown(
                "**🟡 Unrecognised document type** — text preview shown to help diagnose:"
            )
            for msg in unrecognised:
                st.code(msg, language="")

        if image_pdfs:
            names = [f.split(" — ")[0] for f in image_pdfs]
            st.markdown(
                f"**🔵 Scanned / image PDFs ({len(image_pdfs)})** — no selectable text. "
                "Run through OCR (Adobe Acrobat / Tesseract) before uploading."
            )
            with st.expander(f"Show {len(names)} file name(s)"):
                st.write("\n".join(f"• {n}" for n in names))

        if gstr2b_skipped:
            names = [f.split(" — ")[0] for f in gstr2b_skipped]
            st.markdown(
                f"**ℹ️ GSTR-2B files skipped ({len(gstr2b_skipped)})** — "
                "GSTR-2B extraction is not supported in this version."
            )
            with st.expander(f"Show {len(names)} file name(s)"):
                st.write("\n".join(f"• {n}" for n in names))

# ================= OUTPUT =================
has_data = any(st.session_state[t] for t in TYPES)

if has_data:
    st.markdown("---")
    st.subheader("📊 Step 2 — Preview & Download")

    recon_rows = st.session_state.get("recon", [])
    tab_labels = [
        f"🏥 ESI ({len(st.session_state['esi'])})",
        f"🏛 PT ({len(st.session_state['pt'])})",
        f"💰 TDS ({len(st.session_state['tds'])})",
        f"📊 GSTR-1 ({len(st.session_state['gstr1'])})",
        f"📊 GSTR-3B ({len(st.session_state['gstr3b'])})",
        f"🔍 GSTR-1 vs 3B Recon ({len(recon_rows)})",
    ]
    tabs = st.tabs(tab_labels)

    tab_key_map = [
        "esi",
        "pt",
        "tds",
        "gstr1",
        "gstr3b",
    ]

    for tab, key in zip(tabs[:-1], tab_key_map):
        with tab:
            rows = st.session_state[key]
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True)
            else:
                st.info("No records found.")

    with tabs[-1]:
        if recon_rows:
            df_recon = pd.DataFrame(recon_rows)
            # Colour-code by status
            def _highlight(row):
                colour = {"Matched": "#d4edda", "Mismatch": "#f8d7da",
                          "Only in GSTR-1": "#fff3cd", "Only in GSTR-3B": "#cce5ff"}.get(row["Status"], "")
                return [f"background-color: {colour}" if colour else ""] * len(row)
            st.dataframe(df_recon.style.apply(_highlight, axis=1), use_container_width=True)

            # Summary counts
            summary = df_recon["Status"].value_counts().reset_index()
            summary.columns = ["Status", "Count"]
            col_s1, col_s2 = st.columns([1, 2])
            with col_s1:
                st.markdown("**Summary**")
                st.dataframe(summary, use_container_width=True, hide_index=True)
        else:
            st.info("Upload both GSTR-1 and GSTR-3B PDFs to see reconciliation.")

    # ── Excel export ──────────────────────────────────────────────────────────
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_map = [
            ("ESI",    "esi"),
            ("PT",     "pt"),
            ("TDS",    "tds"),
            ("GSTR1",  "gstr1"),
            ("GSTR3B", "gstr3b"),
        ]
        for sheet_name, key in sheet_map:
            rows = st.session_state[key]
            df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["No Data"])
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Reconciliation sheet with conditional formatting
        df_recon_exp = pd.DataFrame(recon_rows) if recon_rows else pd.DataFrame(columns=["No Data"])
        df_recon_exp.to_excel(writer, sheet_name="GSTR1_vs_3B_Recon", index=False)

        if recon_rows:
            from openpyxl.styles import PatternFill, Font
            ws = writer.sheets["GSTR1_vs_3B_Recon"]
            status_col = df_recon_exp.columns.get_loc("Status") + 1  # 1-based
            fill_map = {
                "Matched":         PatternFill("solid", fgColor="D4EDDA"),
                "Mismatch":        PatternFill("solid", fgColor="F8D7DA"),
                "Only in GSTR-1":  PatternFill("solid", fgColor="FFF3CD"),
                "Only in GSTR-3B": PatternFill("solid", fgColor="CCE5FF"),
            }
            diff_cols = [df_recon_exp.columns.get_loc(c) + 1
                         for c in df_recon_exp.columns if c.startswith("Diff")]
            for row_idx, row in enumerate(df_recon_exp.itertuples(index=False), start=2):
                status = getattr(row, "Status", "")
                fill   = fill_map.get(status)
                if fill:
                    for col_idx in range(1, len(df_recon_exp.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
                # Bold non-zero diff cells
                for col_idx in diff_cols:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if float(cell.value or 0) != 0:
                            cell.font = Font(bold=True)
                    except (TypeError, ValueError):
                        pass

    buffer.seek(0)

    st.download_button(
        label="⬇️ Download Excel (All Sheets)",
        data=buffer,
        file_name="Statutory_Compliance_Data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )